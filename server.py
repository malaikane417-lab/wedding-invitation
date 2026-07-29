from flask import Flask, request, jsonify, send_from_directory
import openpyxl
import os
from datetime import datetime

app = Flask(__name__, static_folder='.', static_url_path='')

EXCEL_FILE = os.path.join('rsvp', 'details.xlsx')

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RSVP Details"
        ws.append(["Name", "Email", "Phone", "Guests", "Attending", "Dietary", "Message", "Date"])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22
        wb.save(EXCEL_FILE)

init_excel()

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/src/<path:filename>')
def serve_src(filename):
    return send_from_directory('src', filename)

@app.route('/submit-rsvp', methods=['POST'])
def submit_rsvp():
    data = request.get_json()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        data.get('name', ''),
        data.get('email', ''),
        data.get('phone', ''),
        data.get('guests', ''),
        data.get('attendance', ''),
        data.get('dietary', ''),
        data.get('message', ''),
        datetime.now().strftime('%Y-%m-%d %H:%M')
    ])
    wb.save(EXCEL_FILE)
    return jsonify({'status': 'success', 'message': 'RSVP saved successfully'})

@app.route('/rsvp-data')
def rsvp_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            'name': row[0], 'email': row[1], 'phone': row[2],
            'guests': row[3], 'attendance': row[4], 'dietary': row[5],
            'message': row[6], 'date': row[7]
        })
    return jsonify(rows)

if __name__ == '__main__':
    print("Server running at http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
